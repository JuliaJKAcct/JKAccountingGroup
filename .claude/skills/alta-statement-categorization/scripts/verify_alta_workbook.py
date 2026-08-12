#!/usr/bin/env python3
"""Recalculate an ALTA workbook and enforce every reconciliation.

    python verify_alta_workbook.py --file "Client - ALTA.xlsx"

openpyxl writes formulas but never evaluates them, so a workbook can look right
and still be wrong. This opens it with a real calculation engine and fails loudly
on any check that is not zero. Exits non-zero, so a broken workbook cannot ship.

Checked automatically, wherever the sheet exists:
  * Purchase  K42 / K45 / K48   the template's three reconciliations
  * ANY sheet every "…(should be $0)" row, in column B or F
  * Sale      the disposition journal entry balances (debits = credits)
  * Any sheet a "RESIDUAL" row (the Schedule D / 8949 cash tie-out)
  * Mapping   every tie-out difference in column I

Requires: pip install formulas
"""
import argparse
import re
import sys

try:
    import formulas
except ImportError:
    sys.exit("ERROR: pip install formulas")

import openpyxl

TOL = 0.005


def solve(path):
    """Return {(SHEET_UPPER, 'A1'): value} for every cell in the workbook."""
    solution = formulas.ExcelModel().loads(path).finish().calculate()
    out = {}
    for key, node in solution.items():
        m = re.match(r"^'\[[^\]]+\](.+)'!([A-Z]+\d+)$", key)
        if not m:
            continue
        try:
            out[(m.group(1), m.group(2))] = node.value[0, 0]
        except Exception:
            try:
                out[(m.group(1), m.group(2))] = node.value
            except Exception:
                pass
    return out


def find_rows(ws, col, needle):
    """Row numbers whose cell in `col` contains `needle` (case-insensitive)."""
    hits = []
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, col).value
        if value and needle.lower() in str(value).lower():
            hits.append(row)
    return hits


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--file", required=True)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.file)
    calc = solve(args.file)
    checks, missing = [], []

    def get(sheet, cell):
        return calc.get((sheet.upper(), cell))

    for title in wb.sheetnames:
        ws = wb[title]

        if title.startswith("Purchase - "):
            for cell, what in [("K42", "due from buyer"),
                               ("K45", "paid at closing"),
                               ("K48", "cash to close")]:
                checks.append((f"{title}: {what}", get(title, cell)))

        if title.startswith("Sale - "):
            total = find_rows(ws, 10, "   Total")
            if total:
                row = total[-1]
                debit, credit = get(title, f"K{row}"), get(title, f"L{row}")
                checks.append((f"{title}: journal entry balances",
                               (credit or 0) - (debit or 0)))
                if not debit:
                    missing.append(f"{title}: journal entry debits are empty")
            else:
                missing.append(f"{title}: no journal-entry total row found")

        # Zero-checks are found by their label on EVERY sheet, in whichever
        # column the label sits, rather than by sheet name and a fixed cell -
        # so a new sheet or a renamed one can never silently skip its check.
        for label_col, value_col in ((2, "C"), (6, "G")):
            for row in find_rows(ws, label_col, "should be $0"):
                checks.append((f"{title}: {ws.cell(row, label_col).value.strip()}"
                               f" (row {row})", get(title, f"{value_col}{row}")))
        for row in find_rows(ws, 5, "RESIDUAL"):
            checks.append((f"{title}: cash tie-out residual", get(title, f"F{row}")))

        if title == "ALTA Mapping Chart":
            for row in find_rows(ws, 10, "should be $0"):
                label = ws.cell(row, 2).value or f"row {row}"
                checks.append((f"{title}: {label}", get(title, f"I{row}")))

    if not checks:
        sys.exit("ERROR: nothing to check - are the sheet names as the builder writes them?")

    width = max(len(name) for name, _ in checks)
    failed = 0
    for name, value in checks:
        # An Excel error (#VALUE!, #REF!) is a failure, not a crash - a bad
        # cross-sheet reference shows up here and must not pass silently.
        numeric = isinstance(value, (int, float)) and not isinstance(value, bool)
        bad = not numeric or abs(value) > TOL
        failed += bad
        shown = f"{value:,.2f}" if numeric else str(value)
        print(f"{'FAIL' if bad else 'PASS'}  {name:<{width}}  {shown:>12}")
    for note in missing:
        print(f"WARN  {note}")

    print(f"\n{len(checks) - failed}/{len(checks)} checks passed.")
    if failed:
        print("DO NOT DELIVER until every check is $0.00.")
        return 1
    print("Workbook reconciles. Safe to deliver.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
