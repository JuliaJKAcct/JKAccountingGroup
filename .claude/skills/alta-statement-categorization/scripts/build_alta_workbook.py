#!/usr/bin/env python3
"""Build an ALTA purchase/sale categorization workbook from a deal spec.

    python build_alta_workbook.py --spec deal.json --out "Client - ALTA.xlsx"

Produces up to four sheets:

  1. HUD Tool v3        the blank master template, untouched
  2. Purchase - <tag>   the master prefilled from the purchase ALTA
  3. Sale - <tag>       a disposition worksheet, prefilled from the sale ALTA
  4. ALTA Mapping Chart every ALTA line -> bucket -> target cell, plus a tie-out

The spec addresses template rows by their PRINTED LABEL, never by row number, so
a future edit to the template does not silently write amounts into the wrong row.
An unknown label is a hard error - it never lands in the wrong bucket silently.

Client figures live in the spec and the output workbook. Neither is committed.
"""
import argparse
import datetime as dt
import json
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

NAVY = "FF343F56"
CREAM = "FFF6F6E9"
FLAG = "FF9A3412"
MONEY = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
DATEF = "M/d/yyyy"

fill_navy = PatternFill("solid", fgColor=NAVY)
fill_cream = PatternFill("solid", fgColor=CREAM)

# Where each labelled block lives on the master template: (column of the label,
# first row, last row, column the amount is written to).
PURCHASE_BLOCKS = {
    "facilitative": (2, 18, 45, 3),   # 1 - FACILITATIVE COSTS  -> basis of building
    "loan_costs":   (2, 49, 76, 3),   # 2 - LOAN COSTS          -> basis of loan
    "operating":    (2, 80, 83, 3),   # 3 - OPERATING COSTS     -> expensed
    "escrow":       (2, 87, 90, 3),   # 4 - ESCROW COSTS        -> escrow deposits
    "poc_building": (6, 32, 35, 7),   # paid outside closing, added to building
    "poc_loan":     (6, 39, 42, 7),   # paid outside closing, added to loan
    "reductions":   (6, 46, 53, 7),   # reductions to amount due
}

# The sale worksheet's fixed skeleton. Amounts come from the spec by label.
SELLING_COSTS = [
    "Broker Commission - Listing Agent",
    "Broker Commission - Selling Agent",
    "Broker Transaction | Processing Fee",
    "Title Settlement | Closing Fee",
    "Title | Deed Search Fee",
    "Owner's Title Insurance Premium",
    "Title Endorsements | Doc Preparation",
    "Attorney | Legal Fees",
    "Notary Fees",
    "Municipal Lien | Estoppel Search",
    "Survey Fee",
    "Recording Fees",
    "Govt Recording Fee | E-Recording Fee",
    "Other Recording Documents (Affidavits, TNOC)",
    "Deed Documentary Stamps | Transfer Tax",
    "State | County Tax Stamps",
    "Wire Fee | Courier Fee",
    "Home Warranty | Repairs",
    "Seller Concession | Closing Cost Credit to Buyer",
    "Other",
]
OTHER_DEBITS = [
    "Property Taxes Paid at Closing",
    "Mortgage Payoff - Principal",
    "Mortgage Payoff - Accrued Interest",
    "Other Payoffs | Liens | Judgments",
    "Escrow Holdback Retained by Title",
]
BASIS_LINES = [
    ("Original Contract Purchase Price", "link_price"),
    ("Facilitative Closing Costs Capitalized at Purchase", "link_facil"),
    ("Capital Improvements | Rehab Costs (not on the ALTA)", "input"),
    ("Carrying Costs Capitalized (taxes, insurance, interest)", "input"),
    ("Unamortized Loan Costs Written Off at Sale", "input"),
    ("Less: Accumulated Depreciation Taken", "input"),
]
SELLER_CREDITS = [
    ("Contract Sale Price", "link_price"),
    ("Prorated Property Taxes (Buyer Reimbursement)", "input"),
    ("Prorated Non-Ad Valorem | Special Assessments", "input"),
    ("Prorated Rental Income", "input"),
    ("Tenant Security Deposits Transferred", "input"),
    ("HOA | Condo Dues Proration", "input"),
    ("Other Credits to Seller", "input"),
]
MAP_HEADERS = ["Side", "ALTA Section", "ALTA Line (as printed)", "Column",
               "Bucket", "Target Worksheet", "Target Cell", "This Deal",
               "Tax Treatment"]
MAP_KEYS = ["side", "section", "line", "column", "bucket", "sheet", "cell",
            "amount", "treatment"]


def font(bold=False, cream_text=True, size=10, color=None):
    return Font(name="Montserrat", size=size, bold=bold,
                color=color or (CREAM if cream_text else NAVY))


def dark(cell, bold=False, fmt=None, align=None, wrap=False, size=10):
    """Navy ground, cream ink - the template's 'computed, do not edit' look."""
    cell.fill, cell.font = fill_navy, font(bold, True, size)
    if fmt:
        cell.number_format = fmt
    cell.alignment = Alignment(horizontal=align, wrap_text=wrap, vertical="center")


def light(cell, bold=False, fmt=None, align=None, wrap=False, size=10, color=None):
    """Cream ground, navy ink - the template's 'type here' look."""
    cell.fill, cell.font = fill_cream, font(bold, False, size, color)
    if fmt:
        cell.number_format = fmt
    cell.alignment = Alignment(horizontal=align, wrap_text=wrap, vertical="center")


def as_date(value):
    return dt.datetime.strptime(value, "%Y-%m-%d") if value else None


def resolve(ws, block, label):
    """Find the row whose printed label matches, within one template block."""
    label_col, first, last, amount_col = PURCHASE_BLOCKS[block]
    wanted = label.strip().lower()
    for row in range(first, last + 1):
        cell = ws.cell(row, label_col).value
        if cell and str(cell).strip().lower() == wanted:
            return row, amount_col
    available = [ws.cell(r, label_col).value for r in range(first, last + 1)
                 if ws.cell(r, label_col).value]
    raise SystemExit(
        f"\nERROR: no '{block}' line called {label!r} on the template.\n"
        f"Valid labels are:\n  " + "\n  ".join(str(a) for a in available) + "\n")


# ------------------------------------------------------------------ purchase --
def build_purchase(wb, master, spec, tag):
    pur = wb.copy_worksheet(master)
    pur.title = f"Purchase - {tag}"[:31]
    p = spec["purchase"]

    for cell, value in [
        ("C1", spec.get("entity", "")), ("C2", spec.get("ein", "")),
        ("C3", spec.get("property_street", "")),
        ("C4", spec.get("property_citystatezip", "")),
        ("C5", spec.get("property_type", "")),
        ("C8", p.get("price")),
    ]:
        pur[cell] = value
    pur["C6"] = as_date(p.get("settlement_date"))
    pur["C7"] = as_date(p.get("placed_in_service"))
    pur["C6"].number_format = pur["C7"].number_format = DATEF
    if spec.get("assessor_url"):
        pur["J1"] = spec["assessor_url"]
    if p.get("land_value") is not None:
        pur["C11"] = p["land_value"]
    if p.get("building_value") is not None:
        pur["C12"] = p["building_value"]

    for block in PURCHASE_BLOCKS:
        for label, amount in (p.get(block) or {}).items():
            row, col = resolve(pur, block, label)
            pur.cell(row, col).value = amount

    # Reconciliation pulls, taken straight off the settlement statement.
    for cell, value in (p.get("pull") or {}).items():
        pur[cell] = value

    # Swap the generic download instructions for deal notes.
    for row in range(6, 30):
        pur.cell(row, 14).value = None
    write_notes(pur, p.get("notes", []), col=14)
    return pur


def write_notes(ws, notes, col):
    """Notes are [row, text] or [row, text, bold]; row 0 means 'next line'."""
    cursor = 1
    for note in notes:
        row, text = note[0], note[1]
        bold = note[2] if len(note) > 2 else text.isupper()
        cursor = cursor + 1 if row == 0 else row
        cell = ws.cell(cursor, col)
        cell.value = text
        dark(cell, bold=bold)


# ---------------------------------------------------------------------- sale --
def build_sale(wb, spec, tag, purchase_title):
    sale = wb.create_sheet(f"Sale - {tag}"[:31])
    for col, width in {"A": 3.75, "B": 46.88, "C": 37.38, "D": 4.25, "E": 3.88,
                       "F": 44.5, "G": 15.13, "H": 3.75, "I": 3.75, "J": 42.88,
                       "K": 14.88, "L": 14.88, "M": 12.0, "N": 64.38}.items():
        sale.column_dimensions[col].width = width
    sale.sheet_view.showGridLines = False

    s = spec["sale"]
    amounts = {k.strip().lower(): v for k, v in (s.get("selling_costs") or {}).items()}
    others = {k.strip().lower(): v for k, v in (s.get("other_debits") or {}).items()}
    credits = {k.strip().lower(): v for k, v in (s.get("seller_credits") or {}).items()}
    basis_in = {k.strip().lower(): v for k, v in (s.get("basis") or {}).items()}
    for supplied, valid, what in [(amounts, SELLING_COSTS, "selling_costs"),
                                  (others, OTHER_DEBITS, "other_debits"),
                                  (credits, [c for c, _ in SELLER_CREDITS], "seller_credits"),
                                  (basis_in, [b for b, _ in BASIS_LINES], "basis")]:
        known = {v.strip().lower() for v in valid}
        for key in supplied:
            if key not in known:
                raise SystemExit(
                    f"\nERROR: no '{what}' line called {key!r} on the sale sheet.\n"
                    f"Valid labels are:\n  " + "\n  ".join(valid) + "\n")

    PUR = f"'{purchase_title}'"
    has_purchase = purchase_title is not None

    def label(row, text, col=2, bold=False):
        cell = sale.cell(row, col)
        cell.value = text
        dark(cell, bold=bold)

    def amount(row, value, col=3, computed=False, fmt=MONEY):
        cell = sale.cell(row, col)
        cell.value = value
        (dark if computed else light)(cell, fmt=fmt, align="right")

    def bar(row, first_col, last_col, text):
        cell = sale.cell(row, first_col)
        cell.value = text
        dark(cell, bold=True)
        for col in range(first_col + 1, last_col + 1):
            dark(sale.cell(row, col))
        if last_col > first_col:
            sale.merge_cells(start_row=row, start_column=first_col,
                             end_row=row, end_column=last_col)

    header = [
        ("LLC Name:", spec.get("entity", ""), "@"),
        ("EIN:", spec.get("ein", ""), "@"),
        ("Property Street:", spec.get("property_street", ""), "@"),
        ("Property City, State, Zip", spec.get("property_citystatezip", ""), "@"),
        ("Property Type:", spec.get("property_type", ""), "@"),
        ("Sale (Settlement) Date:", as_date(s.get("settlement_date")), DATEF),
        ("Disbursement Date:", as_date(s.get("disbursement_date")
                                       or s.get("settlement_date")), DATEF),
        ("Contract Sale Price:", s.get("price"), MONEY),
    ]
    for row, (lab, value, fmt) in enumerate(header, start=1):
        label(row, lab)
        amount(row, value, fmt=fmt)

    bar(10, 2, 3, "HOLDING PERIOD")
    purchase_date = (f"={PUR}!C6" if has_purchase
                     else as_date(spec.get("purchase", {}).get("settlement_date")))
    for row, lab, value, fmt in [
        (11, "Original Purchase Date", purchase_date, DATEF),
        (12, "Sale Date", "=C6", DATEF),
        (13, "Days Held", '=IFERROR(C12-C11,"")', "#,##0"),
        (14, "Character of Gain / (Loss)",
         '=IFERROR(IF(C13>365,"Long-Term","Short-Term"),"")', "@"),
    ]:
        label(row, lab)
        amount(row, value, computed=True, fmt=fmt)

    bar(16, 1, 3, "1 - SELLING COSTS - Reduce Amount Realized (Seller Debits)")
    row = 17
    for lab in SELLING_COSTS:
        label(row, lab)
        amount(row, amounts.get(lab.strip().lower(), 0))
        row += 1
    label(row, "TOTAL SELLING COSTS", bold=True)
    amount(row, f"=SUM(C17:C{row - 1})", computed=True)
    r_selling = row

    top = r_selling + 2
    bar(top, 1, 3,
        "2 - OTHER SELLER DEBITS - Expensed or Balance Sheet (Not Selling Costs)")
    row = top + 1
    r_tax, r_prin, r_int, r_lien, r_hold = row, row + 1, row + 2, row + 3, row + 4
    for lab in OTHER_DEBITS:
        label(row, lab)
        amount(row, others.get(lab.strip().lower(), 0))
        row += 1
    label(row, "TOTAL OTHER SELLER DEBITS", bold=True)
    amount(row, f"=SUM(C{top + 1}:C{row - 1})", computed=True)
    r_other = row

    top = r_other + 2
    bar(top, 1, 3, "3 - ADJUSTED BASIS OF PROPERTY SOLD")
    row = top + 1
    r_price, r_facil = row, row + 1
    r_impr, r_carry, r_loanc, r_depr = row + 2, row + 3, row + 4, row + 5
    for lab, kind in BASIS_LINES:
        label(row, lab)
        if kind == "link_price" and has_purchase:
            amount(row, f"={PUR}!C8", computed=True)
        elif kind == "link_facil" and has_purchase:
            amount(row, f"={PUR}!C46", computed=True)
        else:
            amount(row, basis_in.get(lab.strip().lower(), 0))
        row += 1
    label(row, "ADJUSTED BASIS", bold=True)
    amount(row, f"=C{r_price}+C{r_facil}+C{r_impr}+C{r_carry}+C{r_loanc}-C{r_depr}",
           computed=True)
    r_basis = row

    bar(16, 6, 7, "SELLER CREDITS AT CLOSING")
    row = 17
    r_crprice = row
    for lab, kind in SELLER_CREDITS:
        label(row, lab, col=6)
        if kind == "link_price":
            amount(row, "=C8", col=7, computed=True)
        else:
            amount(row, credits.get(lab.strip().lower(), 0), col=7)
        row += 1
    r_crtax, r_crnav = r_crprice + 1, r_crprice + 2
    r_crrent, r_crdep, r_crhoa, r_croth = (r_crprice + 3, r_crprice + 4,
                                           r_crprice + 5, r_crprice + 6)
    label(row, "TOTAL SELLER CREDITS", col=6, bold=True)
    amount(row, f"=SUM(G{r_crprice}:G{row - 1})", col=7, computed=True)
    r_credits = row

    bar(row + 2, 6, 7, "PROCEEDS RECONCILIATION")
    row += 3
    for lab, value in [("Total Seller Credits", f"=G{r_credits}"),
                       ("Less: Total Selling Costs", f"=-C{r_selling}"),
                       ("Less: Total Other Seller Debits", f"=-C{r_other}")]:
        label(row, lab, col=6)
        amount(row, value, col=7, computed=True)
        row += 1
    label(row, "NET DUE TO SELLER (Calculated)", col=6, bold=True)
    amount(row, f"=SUM(G{r_credits + 3}:G{row - 1})", col=7, computed=True)
    r_netdue = row
    label(row + 1, "   Pull from Settlement Statement", col=6, bold=True)
    amount(row + 1, (s.get("pull") or {}).get("net_due_to_seller"), col=7)
    label(row + 2, "Reconciliation (should be $0)", col=6)
    amount(row + 2, f"=ROUND(G{r_netdue}-G{row + 1},2)", col=7, computed=True)

    row += 4
    label(row, "Total Seller Debits (Calculated)", col=6, bold=True)
    amount(row, f"=C{r_selling}+C{r_other}", col=7, computed=True)
    label(row + 1, "   Pull from Settlement Statement", col=6, bold=True)
    amount(row + 1, (s.get("pull") or {}).get("total_seller_debits"), col=7)
    label(row + 2, "Reconciliation (should be $0)", col=6)
    amount(row + 2, f"=ROUND(G{row}-G{row + 1},2)", col=7, computed=True)

    bar(16, 9, 11, "GAIN / (LOSS) ON SALE")
    for row, lab, value in [
        (17, "Contract Sale Price", "=C8"),
        (18, "Less: Selling Costs", f"=-C{r_selling}"),
        (19, "AMOUNT REALIZED", "=K17+K18"),
        (20, "Less: Adjusted Basis", f"=-C{r_basis}"),
        (21, "TOTAL GAIN / (LOSS)", "=K19+K20"),
    ]:
        label(row, lab, col=10, bold=lab.isupper())
        amount(row, value, col=11, computed=True)
    r_gain = 21
    for row, lab, value, fmt in [
        (23, "Character (holding period)", "=C14", "@"),
        (24, "Unrecaptured Sec. 1250 Gain (Depreciation Recapture)",
         f"=IF(K{r_gain}>0,MIN(K{r_gain},C{r_depr}),0)", MONEY),
        (25, "Remaining Sec. 1231 / Capital Gain or (Loss)", f"=K{r_gain}-K24", MONEY),
    ]:
        label(row, lab, col=10)
        amount(row, value, col=11, computed=True, fmt=fmt)

    dark(sale.cell(28, 9))
    label(28, "Journal Entry - Disposition", col=10, bold=True)
    for col, text in ((11, "Debit"), (12, "Credit")):
        cell = sale.cell(28, col)
        cell.value = text
        dark(cell, bold=True, align="center")

    row = 29
    for lab, value in [
        ("Cash / Bank (Net Proceeds Received)", f"=G{r_netdue}"),
        ("Escrow Receivable - Holdback (Current Asset)", f"=C{r_hold}"),
        ("Accumulated Depreciation (contra Fixed Asset)", f"=C{r_depr}"),
        ("Property Taxes Paid (Expense)", f"=C{r_tax}"),
        ("Real Estate Mortgage (Long Term Liability)", f"=C{r_prin}"),
        ("Interest Paid (Expense)", f"=C{r_int}"),
        ("Other Payoffs | Liens (Expense or Liability)", f"=C{r_lien}"),
        ("Loss on Sale of Property", f"=IF(K{r_gain}<0,-K{r_gain},0)"),
    ]:
        label(row, lab, col=10)
        amount(row, value, col=11, computed=True)
        dark(sale.cell(row, 12), fmt=MONEY)
        row += 1
    r_deb_end = row - 1
    for lab, value in [
        ("Building | Land | Improvements (Fixed Asset)",
         f"=C{r_price}+C{r_facil}+C{r_impr}+C{r_carry}"),
        ("Loan Costs (Fixed Asset)", f"=C{r_loanc}"),
        ("Property Taxes Paid (Expense) - Buyer Reimbursement",
         f"=G{r_crtax}+G{r_crnav}"),
        ("Rental Income (Income)", f"=G{r_crrent}"),
        ("Tenant Security Deposits (Current Liability)", f"=G{r_crdep}"),
        ("HOA | Condo Fees (Expense) - Buyer Reimbursement", f"=G{r_crhoa}"),
        ("Other Credits to Seller", f"=G{r_croth}"),
        ("Gain on Sale of Property", f"=IF(K{r_gain}>0,K{r_gain},0)"),
    ]:
        label(row, lab, col=10)
        dark(sale.cell(row, 11), fmt=MONEY)
        amount(row, value, col=12, computed=True)
        row += 1
    r_cre_end = row - 1

    row += 1
    label(row, "   Total", col=10, bold=True)
    amount(row, f"=SUM(K29:K{r_deb_end})", col=11, computed=True)
    amount(row, f"=SUM(L{r_deb_end + 1}:L{r_cre_end})", col=12, computed=True)
    amount(row, f"=L{row}-K{row}", col=13, computed=True)

    write_notes(sale, s.get("notes", []), col=14)
    for row in range(1, r_basis + 6):
        for col in range(1, 15):
            cell = sale.cell(row, col)
            if cell.fill.patternType is None:
                cell.fill = fill_cream
    return sale


# ------------------------------------------------------------------- mapping --
def build_mapping(wb, spec):
    ws = wb.create_sheet("ALTA Mapping Chart")
    ws.sheet_view.showGridLines = False
    for col, width in {"A": 3.0, "B": 11.0, "C": 27.0, "D": 44.0, "E": 9.0,
                       "F": 13.0, "G": 34.0, "H": 20.0, "I": 12.0, "J": 14.0,
                       "K": 62.0}.items():
        ws.column_dimensions[col].width = width

    ws["B1"] = "ALTA SETTLEMENT STATEMENT - CATEGORIZATION CHART"
    dark(ws["B1"], bold=True)
    for col in range(3, 12):
        dark(ws.cell(1, col))
    ws.merge_cells("B1:K1")
    ws["B2"] = ("Every line on the statement(s), in printed order. 'Column' is whose "
                "side of the ALTA the amount sits on - the other party's lines never "
                "touch our client's books.")
    light(ws["B2"], wrap=True, size=8)
    for col in range(3, 12):
        light(ws.cell(2, col), size=8)
    ws.merge_cells("B2:K2")
    ws.row_dimensions[2].height = 26

    for i, head in enumerate(MAP_HEADERS):
        cell = ws.cell(4, 2 + i)
        cell.value = head
        dark(cell, bold=True, align="center", wrap=True)
    ws.row_dimensions[4].height = 30

    row = 5
    for entry in spec.get("mapping", []):
        for i, key in enumerate(MAP_KEYS):
            cell = ws.cell(row, 2 + i)
            cell.value = entry.get(key, "")
            excluded = key == "bucket" and "exclude" in str(entry.get(key, "")).lower()
            light(cell, wrap=key in ("line", "treatment"), size=9,
                  bold=excluded, color=FLAG if excluded else None)
            if key == "amount":
                cell.number_format = MONEY
                cell.alignment = Alignment(horizontal="right", vertical="center")
        ws.row_dimensions[row].height = 26
        row += 1

    tie = spec.get("tie_out", [])
    if tie:
        row += 1
        ws.cell(row, 2).value = "TIE-OUT TO THE STATEMENTS"
        dark(ws.cell(row, 2), bold=True)
        for col in range(3, 12):
            dark(ws.cell(row, col))
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=11)
        row += 1
        for entry in tie:
            ws.cell(row, 2).value = entry["label"]
            for col in (2, 3, 4):
                light(ws.cell(row, col), size=9)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            ws.cell(row, 5).value = "Per ALTA"
            light(ws.cell(row, 5), size=9, align="right")
            ws.cell(row, 6).value = entry["per_alta"]
            light(ws.cell(row, 6), fmt=MONEY, size=9, align="right")
            ws.cell(row, 7).value = "Per workbook"
            light(ws.cell(row, 7), size=9, align="right")
            ws.cell(row, 8).value = entry["formula"]
            dark(ws.cell(row, 8), fmt=MONEY, size=9, align="right")
            ws.cell(row, 9).value = f"=ROUND(F{row}-H{row},2)"
            dark(ws.cell(row, 9), fmt=MONEY, size=9, align="right")
            ws.cell(row, 10).value = "difference (should be $0)"
            light(ws.cell(row, 10), size=9)
            light(ws.cell(row, 11), size=9)
            row += 1

    for r in range(1, row + 2):
        for col in range(1, 13):
            cell = ws.cell(r, col)
            if cell.fill.patternType is None:
                cell.fill = fill_cream
    return ws


# ------------------------------------------------------- carrying costs (COGS) --
def build_carrying(wb, spec, tag):
    """Tiered schedule of post-closing costs capitalized into basis.

    A property that never went into service has no operating expenses: every
    cost from acquisition to disposition is a cost of the property. This sheet
    grades them by how well they support that, so the preparer chooses a tier
    rather than inheriting one silently.
    """
    ws = wb.create_sheet(f"Carrying Costs - {tag}"[:31])
    ws.sheet_view.showGridLines = False
    for col, width in {"A": 3.0, "B": 12.0, "C": 40.0, "D": 30.0, "E": 15.0,
                       "F": 3.0, "G": 46.0, "H": 15.0}.items():
        ws.column_dimensions[col].width = width

    cc = spec["carrying_costs"]
    ws["B1"] = f"CARRYING COSTS - {cc.get('source', 'per the bank ledger')}"
    dark(ws["B1"], bold=True)
    for col in range(3, 6):
        dark(ws.cell(1, col))
    ws.merge_cells("B1:E1")
    ws["B2"] = cc.get("intro", "")
    light(ws["B2"], wrap=True, size=8)
    for col in range(3, 6):
        light(ws.cell(2, col), size=8)
    ws.merge_cells("B2:E2")
    ws.row_dimensions[2].height = 28

    row = 4
    subtotal_rows = {}
    for tier_name in sorted(cc["tiers"]):
        entries = cc["tiers"][tier_name]
        ws.cell(row, 2).value = tier_name
        dark(ws.cell(row, 2), bold=True)
        for col in range(3, 6):
            dark(ws.cell(row, col))
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        row += 1
        for head in (["Date", "Payee", "Ledger category", "Amount"],):
            for i, text in enumerate(head):
                cell = ws.cell(row, 2 + i)
                cell.value = text
                light(cell, bold=True, size=8,
                      align="right" if text == "Amount" else None)
        row += 1
        first = row
        for entry in entries:
            for i, value in enumerate(entry):
                cell = ws.cell(row, 2 + i)
                cell.value = value
                light(cell, size=9, align="right" if i == 3 else None)
                if i == 3:
                    cell.number_format = MONEY
            row += 1
        ws.cell(row, 2).value = f"Subtotal - {tier_name.split('.')[0]}"
        dark(ws.cell(row, 2), bold=True)
        for col in (3, 4):
            dark(ws.cell(row, col))
        ws.cell(row, 5).value = f"=SUM(E{first}:E{row - 1})"
        dark(ws.cell(row, 5), fmt=MONEY, align="right", bold=True)
        subtotal_rows[tier_name] = row
        row += 2

    ws.cell(row, 7).value = "WHICH TIERS GO INTO BASIS"
    dark(ws.cell(row, 7), bold=True)
    dark(ws.cell(row, 8))
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
    row += 1
    included = set(cc.get("include", []))
    first_pick = row
    for tier_name in sorted(cc["tiers"]):
        key = tier_name.split(".")[0]
        ws.cell(row, 7).value = f"{tier_name}   {'INCLUDED' if key in included else 'excluded'}"
        light(ws.cell(row, 7), size=9, bold=key in included)
        ws.cell(row, 8).value = (f"=E{subtotal_rows[tier_name]}" if key in included else 0)
        light(ws.cell(row, 8), fmt=MONEY, align="right", size=9)
        row += 1
    ws.cell(row, 7).value = "TOTAL CAPITALIZED INTO BASIS"
    dark(ws.cell(row, 7), bold=True)
    ws.cell(row, 8).value = f"=SUM(H{first_pick}:H{row - 1})"
    dark(ws.cell(row, 8), fmt=MONEY, align="right", bold=True)
    ws.carrying_total_ref = f"'{ws.title}'!H{row}"

    for r in range(1, row + 2):
        for col in range(1, 10):
            if ws.cell(r, col).fill.patternType is None:
                ws.cell(r, col).fill = fill_cream
    return ws


# ---------------------------------------------------------------- schedule D --
def build_schedule_d(wb, spec, tag, carrying_ref):
    """The filing-ready gain/loss, with the cash tie-out that proves it."""
    d_early = spec["schedule_d"]
    ws = wb.create_sheet(d_early.get("sheet_name", "Form 8949 - Schedule D")[:31])
    ws.sheet_view.showGridLines = False
    for col, width in {"A": 3.0, "B": 52.0, "C": 18.0, "D": 4.0,
                       "E": 52.0, "F": 18.0}.items():
        ws.column_dimensions[col].width = width
    d = spec["schedule_d"]

    def block(row, title, cols=(2, 3)):
        ws.cell(row, cols[0]).value = title
        dark(ws.cell(row, cols[0]), bold=True)
        dark(ws.cell(row, cols[1]))
        ws.merge_cells(start_row=row, start_column=cols[0],
                       end_row=row, end_column=cols[1])

    def line(row, text, value, col=2, computed=False, bold=False):
        ws.cell(row, col).value = text
        dark(ws.cell(row, col), bold=bold)
        cell = ws.cell(row, col + 1)
        cell.value = value
        (dark if computed else light)(cell, fmt=MONEY, align="right", bold=bold)

    ws["B1"] = d.get("heading", f"FORM 8949 / SCHEDULE D - {tag}")
    dark(ws["B1"], bold=True)
    dark(ws["C1"])
    ws.merge_cells("B1:C1")
    ws["B2"] = d.get("intro", "")
    light(ws["B2"], wrap=True, size=8)
    light(ws["C2"], size=8)
    ws.merge_cells("B2:C2")
    ws.row_dimensions[2].height = 40

    block(4, "AMOUNT REALIZED")
    line(5, "Contract Sale Price", d["sale_price"])
    line(6, "Less: Selling Costs (per the sale ALTA)", -abs(d["selling_costs"]))
    line(7, "AMOUNT REALIZED", "=C5+C6", computed=True, bold=True)

    block(9, "ADJUSTED BASIS")
    line(10, "Purchase closing - total paid to title", d["purchase_closing"])
    line(11, "Property taxes paid at the sale closing", d["property_tax"])
    line(12, "Less: buyer's reimbursement of those taxes", -abs(d["tax_reimbursement"]))
    line(13, "Carrying costs capitalized (see Carrying Costs)",
         f"={carrying_ref}" if carrying_ref else d.get("carrying_costs", 0),
         computed=bool(carrying_ref))
    line(14, "Capital improvements | rehab (add if any)", d.get("improvements", 0))
    line(15, "ADJUSTED BASIS", "=SUM(C10:C14)", computed=True, bold=True)

    block(17, "GAIN / (LOSS)")
    line(18, "Amount Realized", "=C7", computed=True)
    line(19, "Less: Adjusted Basis", "=-C15", computed=True)
    line(20, "CAPITAL GAIN / (LOSS)", "=C18+C19", computed=True, bold=True)
    ws.cell(21, 2).value = "Holding period"
    dark(ws.cell(21, 2))
    ws.cell(21, 3).value = d.get("holding", "")
    dark(ws.cell(21, 3), align="right")

    block(4, "CASH TIE-OUT - proves nothing is missing or double counted", (5, 6))
    for row, text, value, computed in [
        (5, "Proceeds actually received", d["cash_received"], False),
        (6, "Less: total paid to title at purchase", -abs(d["purchase_closing"]), False),
        # Reads C13 rather than the carrying sheet again: one source, and a
        # leading unary minus on a cross-sheet reference does not evaluate
        # everywhere.
        (7, "Less: carrying costs paid from the bank", "=-C13", True),
        (8, "NET CASH OUT OF THE DEAL", "=SUM(F5:F7)", True),
        (10, "Schedule D gain / (loss)", "=C20", True),
        (11, "Difference", "=F8-F10", True),
        (13, "Escrow holdback still receivable", -abs(d.get("escrow_holdback", 0)), False),
        (14, "Proceeds short of the ALTA (unexplained)",
         -abs(d.get("proceeds_shortfall", 0)), False),
        (15, "RESIDUAL - must be $0.00", "=ROUND(F11-F13-F14,2)", True),
    ]:
        line(row, text, value, col=5, computed=computed,
             bold=text.isupper() or "RESIDUAL" in text)

    for note_row, text in enumerate(d.get("notes", []), start=17):
        ws.cell(note_row, 5).value = text
        light(ws.cell(note_row, 5), wrap=True, size=9,
              bold=text.isupper() or text.endswith("?"))
        light(ws.cell(note_row, 6), size=9)
        ws.merge_cells(start_row=note_row, start_column=5,
                       end_row=note_row, end_column=6)
        ws.row_dimensions[note_row].height = 26

    for r in range(1, 40):
        for col in range(1, 8):
            if ws.cell(r, col).fill.patternType is None:
                ws.cell(r, col).fill = fill_cream
    return ws


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--spec", required=True, help="deal spec JSON")
    ap.add_argument("--out", required=True, help="output .xlsx")
    ap.add_argument("--template", default=None, help="blank HUD template .xlsx")
    args = ap.parse_args()

    spec = json.load(open(args.spec))
    template = args.template or (
        __file__.rsplit("/scripts/", 1)[0] + "/assets/HUD-Tool-v3-blank.xlsx")

    wb = openpyxl.load_workbook(template)
    master = wb["HUD Tool v3"]
    tag = spec.get("tag", "Property")
    order = [master]

    purchase_title = None
    if spec.get("purchase"):
        pur = build_purchase(wb, master, spec, tag)
        purchase_title = pur.title
        order.append(pur)
    if spec.get("sale"):
        order.append(build_sale(wb, spec, tag, purchase_title))
    carrying_ref = None
    if spec.get("carrying_costs"):
        carrying = build_carrying(wb, spec, tag)
        carrying_ref = carrying.carrying_total_ref
        order.append(carrying)
    if spec.get("schedule_d"):
        order.append(build_schedule_d(wb, spec, tag, carrying_ref))
    if spec.get("mapping") or spec.get("tie_out"):
        order.append(build_mapping(wb, spec))

    wb._sheets = order
    wb.active = 1 if len(order) > 1 else 0
    wb.save(args.out)
    print(f"wrote {args.out}")
    print("sheets: " + ", ".join(s.title for s in order))
    print("\nNow VERIFY before delivering (see SKILL.md step 5):")
    print("  python " + __file__.rsplit("/", 1)[0] + "/verify_alta_workbook.py --file "
          + args.out)
    return 0


if __name__ == "__main__":
    sys.exit(main())
