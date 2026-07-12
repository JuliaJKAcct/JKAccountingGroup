#!/usr/bin/env python3
"""
build_expenses_report.py

Turn a QuickBooks **Transaction Detail by Account** export (.xlsx) into a client
**Expenses report** that ties to the Profit & Loss, by deleting the account
sections that are not expenses (bank, A/R, A/P, credit cards, other current
liabilities, equity/Distributions, income/Sales). The original QuickBooks
formatting (column widths, bold totals, currency format, footer) is preserved —
this only removes whole account sections and fixes the grand TOTAL.

Why a script: the raw export uses live formulas for every subtotal
(``=H7+H8+...``) and the grand total. If you just delete rows in Excel, those
formulas silently point at the wrong cells. This tool first freezes every
formula to its current value, then deletes the sections, then re-derives the
grand TOTAL from the surviving transaction rows — so nothing can drift.

No client data lives in this file. You pass the account labels to remove and,
ideally, the P&L expense total to verify against.

Usage
-----
    python build_expenses_report.py \
        --src  "Transaction Detail by Account.xlsx" \
        --out  "Expenses by Account.xlsx" \
        --remove "Example Co (1234) - 1" "Accounts Receivable (A/R)" \
                 "J. Doe (5678) - 1" "Used personal card" \
                 "Distributions" "Sales" \
        --expect-total 12345.67

`--remove` takes the EXACT text of each non-expense account header as it appears
in column A of the export. `--expect-total` is the "Total for Expenses" figure
from the P&L (same period, same basis); the script exits non-zero if the cleaned
file does not tie to it, so a wrong/missed section can never ship silently.
"""
import argparse
import sys

import openpyxl

AMOUNT_COL = 8  # column H in a QuickBooks Transaction Detail by Account export


def col_a(ws, r):
    return ws.cell(row=r, column=1).value


def amount(ws, r):
    return ws.cell(row=r, column=AMOUNT_COL).value


def find_block(ws, label):
    """Inclusive (start, end) row range of the top-level account section whose
    header (column A) equals ``label``. The section closes at the first row whose
    column A is ``Total for {label}`` or ``Total for {label} with sub-accounts``
    — which correctly skips over any nested sub-account totals in between."""
    start = next((r for r in range(1, ws.max_row + 1) if col_a(ws, r) == label), None)
    if start is None:
        sys.exit(f"ERROR: account section not found in column A: {label!r}")
    closers = {f"Total for {label}", f"Total for {label} with sub-accounts"}
    for r in range(start + 1, ws.max_row + 1):
        if col_a(ws, r) in closers:
            return start, r
    sys.exit(f"ERROR: no closing 'Total for' row found for section {label!r}")


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--src", required=True, help="Source .xlsx (Transaction Detail by Account)")
    ap.add_argument("--out", required=True, help="Destination .xlsx (expenses-only)")
    ap.add_argument("--remove", nargs="+", required=True,
                    help="Exact column-A labels of the non-expense sections to delete")
    ap.add_argument("--expect-total", type=float, default=None,
                    help="P&L total expenses; the script fails if the result does not tie to it")
    ap.add_argument("--tol", type=float, default=0.005, help="Rounding tolerance (default 0.005)")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.src, data_only=False)   # editable: formatting + formulas
    vals = openpyxl.load_workbook(args.src, data_only=True)  # cached formula results
    ws, vs = wb.active, vals.active

    # 1) Freeze every formula to its current value so row deletion can't break subtotals.
    missing_cache = False
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                cached = vs[c.coordinate].value
                if cached is None:
                    missing_cache = True
                c.value = cached
    if missing_cache:
        print("WARNING: the source has formulas with no cached values, so some subtotals "
              "may come out blank. Open and re-save the export in Excel, then re-run.",
              file=sys.stderr)

    # 2) Detach a full-width footer note (e.g. "Accrual Basis <timestamp>") before rows shift.
    footer_text = None
    for m in list(ws.merged_cells.ranges):
        txt = ws.cell(row=m.min_row, column=1).value
        if isinstance(txt, str) and (txt.startswith("Accrual Basis") or txt.startswith("Cash Basis")):
            footer_text = txt
            ws.unmerge_cells(str(m))
            break

    # 3) Delete each non-expense section, bottom-up so earlier row numbers stay valid.
    blocks = sorted((find_block(ws, lbl) for lbl in args.remove),
                    key=lambda b: b[0], reverse=True)
    for start, end in blocks:
        ws.delete_rows(start, end - start + 1)

    # 4) Re-derive the grand TOTAL from the surviving transaction (leaf) rows.
    #    Leaf rows have an empty column A and a numeric amount; header and
    #    "Total for ..." rows are excluded, so this equals the true expenses total.
    grand = round(sum(
        amount(ws, r) for r in range(1, ws.max_row + 1)
        if (col_a(ws, r) in (None, "")) and isinstance(amount(ws, r), (int, float))
    ), 2)
    total_written = False
    for r in range(1, ws.max_row + 1):
        if col_a(ws, r) == "TOTAL":
            ws.cell(row=r, column=AMOUNT_COL).value = grand
            total_written = True
    if not total_written:
        sys.exit("ERROR: no grand-total row (column A == 'TOTAL') found, so the tie-out "
                 "total could not be set. Check the export's grand-total label.")

    # 5) Re-place the footer on its new row (if present) and trim phantom trailing rows
    #    (openpyxl's delete_rows can leave empty rows past the real content behind).
    if footer_text is not None:
        frow = next((r for r in range(1, ws.max_row + 1) if col_a(ws, r) == footer_text), None)
        if frow:
            ws.merge_cells(start_row=frow, start_column=1, end_row=frow, end_column=8)
    last = max((r for r in range(1, ws.max_row + 1)
                if any(ws.cell(row=r, column=c).value not in (None, "") for c in range(1, 9))),
               default=ws.max_row)
    if ws.max_row > last:
        ws.delete_rows(last + 1, ws.max_row - last)

    wb.save(args.out)

    # 6) Report the surviving section subtotals and verify the tie-out against the
    #    total actually written to the saved file (not just the computed value).
    out = openpyxl.load_workbook(args.out, data_only=True).active
    saved_total = next((out.cell(row=r, column=AMOUNT_COL).value
                        for r in range(1, out.max_row + 1)
                        if out.cell(row=r, column=1).value == "TOTAL"), None)
    print(f"Saved: {args.out}")
    print("Surviving section subtotals:")
    for r in range(1, out.max_row + 1):
        a = out.cell(row=r, column=1).value
        if isinstance(a, str) and a.startswith("Total for"):
            print(f"  {a:<48} {out.cell(row=r, column=AMOUNT_COL).value:,.2f}")
    print(f"\nGrand TOTAL (in file) = {saved_total:,.2f}")
    if args.expect_total is not None:
        if saved_total is None or abs(saved_total - args.expect_total) > args.tol:
            sys.exit(f"MISMATCH: file total {saved_total} != P&L {args.expect_total:,.2f}. "
                     "Do NOT send this file — re-check which sections you removed.")
        print(f"OK: ties to the P&L total of {args.expect_total:,.2f}")


if __name__ == "__main__":
    main()
