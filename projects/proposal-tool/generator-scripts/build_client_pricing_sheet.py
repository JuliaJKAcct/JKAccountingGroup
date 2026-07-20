# Build the INTERNAL, client-specific monthly-pricing sheet on top of the shared
# "Rate Tables" reference sheet produced by build_pricing_xlsx.py.
#
# This is the pricing engine realized as a spreadsheet: it computes the bundled
# monthly fee from a handful of client inputs, using live formulas that reference
# Rate Tables, then applies a single negotiated adjustment to reach the final
# client-facing price. It is INTERNAL ONLY — the client proposal never itemizes
# this; it shows one bundled monthly fee (see ../docs/methodology.md, "Bundling
# rule"). Recalc with the xlsx skill's recalc.py after any edit to confirm zero
# formula errors.
#
# Usage: this is a TEMPLATE. Replace every blue input cell (the CLIENT INPUTS
# block and the Negotiated Adjustment) with the real client's figures, set
# CLIENT_NAME, and run. It reads /tmp/pricing_wip.xlsx (written by
# build_pricing_xlsx.py) and writes ./output/<client>_Internal_Pricing_Breakdown.xlsx.

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment

TEAL = "123841"
BRONZE = "9C6A39"
IVORY = "F6F3EC"
WHITE = "FFFFFF"
BLUE = "0000FF"
DARKEST = "091F24"

CLIENT_NAME = "[Client Name]"           # <-- set to the client / entity name
SHEET_NAME = "Client - Pricing"

wb = load_workbook("/tmp/pricing_wip.xlsx")

ws = wb.create_sheet(SHEET_NAME)
RT = "'Rate Tables'"

def header(cell, text, span=None):
    ws[cell] = text
    ws[cell].font = Font(bold=True, color=WHITE, name="Arial", size=11)
    ws[cell].fill = PatternFill("solid", fgColor=TEAL)
    if span:
        ws.merge_cells(span)

def label(cell, text, bold=False):
    ws[cell] = text
    ws[cell].font = Font(bold=bold, name="Arial")

def blue(cell, value):
    ws[cell] = value
    ws[cell].font = Font(color=BLUE, name="Arial")

def calc(cell, formula):
    ws[cell] = formula
    ws[cell].font = Font(color="000000", name="Arial")

ws["A1"] = f"{CLIENT_NAME} — Internal Monthly Pricing Breakdown (NOT for client use)"
ws["A1"].font = Font(bold=True, size=13, color=TEAL, name="Arial")
ws.merge_cells("A1:D1")
ws["A2"] = "Internal calculation only. Blue cells are inputs — replace with the client's real figures; every other cell is a live formula. The client proposal shows ONE bundled monthly fee, never this itemization."
ws["A2"].font = Font(italic=True, size=9, color="6F6857", name="Arial")
ws.merge_cells("A2:F2")

# --- Inputs section (blue = replace per client) ---
header("A4", "CLIENT INPUTS", span="A4:B4")
inputs = [
    ("Average transactions per month", 200, "A5"),
    ("Recording frequency", "Monthly", "A6"),
    ("Bank feeds connected?", "Yes", "A7"),
    ("Number of business locations", 1, "A8"),
    ("Number of bank/CC accounts to reconcile", 5, "A9"),
    ("Accounting Advisory Services tier", "Small", "A10"),
    ("Annual tax prep fee (by return type)", 750, "A11"),
    ("Sales tax filing frequency", "Monthly", "A12"),
    ("Number of states for sales tax", 1, "A13"),
    ("Number of 1099 contractors", 0, "A14"),
    ("Owner Payroll — solo S-corp owner?", "Yes", "A15"),
]
for i,(lbl,val,cell) in enumerate(inputs):
    r = 5+i
    label(f"A{r}", lbl)
    blue(f"B{r}", val)

# --- Calculated monthly fee build-up ---
header("A17", "MONTHLY FEE CALCULATION", span="A17:D17")
header("A18", "Component")
header("B18", "Formula basis")
header("C18", "Monthly $")

# Recording Accounting Transactions
label("A19", "Recording Accounting Transactions")
label("B19", "$1 x tier value x frequency mult x bank-feed mult x location mult")
calc("C19", f"=1*LOOKUP(B5,{RT}!$A$5:$A$16,{RT}!$C$5:$C$16)*VLOOKUP(B6,{RT}!$A$22:$B$24,2,FALSE)*VLOOKUP(B7,{RT}!$A$28:$B$29,2,FALSE)*VLOOKUP(B8,{RT}!$A$33:$B$39,2,FALSE)")

# Reconcile accounts
label("A20", "Reconcile Bank/CC/PayPal Accounts")
label("B20", "$18.50 x account tier value")
calc("C20", f"=18.5*LOOKUP(B9,{RT}!$A$43:$A$47,{RT}!$C$43:$C$47)")

# Financial statements
label("A21", "Prepare Financial Statements (P&L + Balance Sheet)")
label("B21", "Flat")
calc("C21", f"={RT}!$B$70")

# Tax prep
label("A22", "Tax Prep (by return type), amortized monthly")
label("B22", "Annual fee / 12")
calc("C22", "=B11/12")

# Sales tax
label("A23", "Sales Tax Filing")
label("B23", "$60 x states / filing-frequency divisor")
calc("C23", f"=60*B13/VLOOKUP(B12,{RT}!$A$59:$B$61,2,FALSE)")

# Advisory
label("A24", "Accounting Advisory Services")
label("B24", "Flat by company size tier")
calc("C24", f"=VLOOKUP(B10,{RT}!$A$53:$B$55,2,FALSE)")

# 1099
label("A25", "1099 Filing")
label("B25", "$25 x number of contractors")
calc("C25", f"=B14*{RT}!$B$69")

# Owner payroll
label("A26", "Owner Payroll (complimentary if solo S-corp owner)")
label("B26", "$0 if solo S-corp owner, else N/A — not part of this bundle")
calc("C26", '=IF(B15="Yes",0,"N/A")')

# Calculated total (internal matrix, before any negotiated adjustment)
ws["A27"] = "CALCULATED TOTAL (internal matrix)"
ws["A27"].font = Font(bold=True, color=WHITE, name="Arial", size=11)
ws["A27"].fill = PatternFill("solid", fgColor=TEAL)
ws.merge_cells("A27:B27")
calc("C27", "=SUM(C19:C25)")
ws["C27"].font = Font(bold=True, color=WHITE, name="Arial", size=11)
ws["C27"].fill = PatternFill("solid", fgColor=TEAL)

# Negotiated adjustment (blue input) + final client-facing price (formula)
label("A28", "Negotiated Adjustment (client-agreed discount)")
blue("B28", 0)
ws["B28"].number_format = '$#,##0.00;($#,##0.00);-'
calc("C28", "=B28")

ws["A29"] = "FINAL MONTHLY PRICE (client-facing)"
ws["A29"].font = Font(bold=True, color=WHITE, name="Arial", size=11)
ws["A29"].fill = PatternFill("solid", fgColor=DARKEST)
ws.merge_cells("A29:B29")
calc("C29", "=C27+C28")
ws["C29"].font = Font(bold=True, color=WHITE, name="Arial", size=11)
ws["C29"].fill = PatternFill("solid", fgColor=DARKEST)

for cell in ["C19","C20","C21","C22","C23","C24","C25","C27","C28","C29"]:
    ws[cell].number_format = '$#,##0.00'

ws["A31"] = "Note: internal calculation only. The client-facing proposal never itemizes this — it shows a single bundled monthly fee (see ../docs/methodology.md). The Negotiated Adjustment row is the single discount off the bundled total (enter as a negative number)."
ws["A31"].font = Font(italic=True, size=9, color="6F6857", name="Arial")
ws.merge_cells("A31:F31")

ws.column_dimensions["A"].width = 46
ws.column_dimensions["B"].width = 58
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 12

# reorder so client sheet is first
wb.move_sheet(SHEET_NAME, offset=-1)
wb.active = 0

os.makedirs("output", exist_ok=True)
slug = CLIENT_NAME.strip("[]").replace(" ", "_") or "Client"
out = os.path.join("output", f"{slug}_Internal_Pricing_Breakdown.xlsx")
wb.save(out)
print("saved", out)
