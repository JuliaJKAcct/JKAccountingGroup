from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TEAL = "123841"
BRONZE = "9C6A39"
IVORY = "F6F3EC"
WHITE = "FFFFFF"
BLUE = "0000FF"
BLACK = "000000"

wb = Workbook()

# ---------- Sheet 1: Rate Tables (reference, reusable across clients) ----------
rt = wb.active
rt.title = "Rate Tables"

def header(ws, cell, text, width=None):
    ws[cell] = text
    ws[cell].font = Font(bold=True, color=WHITE, name="Arial")
    ws[cell].fill = PatternFill("solid", fgColor=TEAL)

def note(ws, cell, text):
    ws[cell] = text
    ws[cell].font = Font(italic=True, size=9, color="6F6857", name="Arial")

rt["A1"] = "JKA Core Pricing Matrix — Rate Tables (internal reference, applies to all clients)"
rt["A1"].font = Font(bold=True, size=13, color=TEAL, name="Arial")
rt.merge_cells("A1:D1")

# Transactions/month tier table
rt["A3"] = "Recording Transactions — Monthly Volume Tiers"
rt["A3"].font = Font(bold=True, name="Arial")
header(rt, "A4", "Lower Bound (>=)")
header(rt, "B4", "Upper Bound")
header(rt, "C4", "Billed Tier Value")
tiers = [(0,200,200),(201,225,225),(226,250,250),(251,275,275),(276,300,300),
         (301,325,325),(326,350,350),(351,375,375),(376,400,400),(401,425,425),
         (426,450,450),(451,500,500)]
for i,(lo,hi,val) in enumerate(tiers):
    r = 5+i
    rt[f"A{r}"] = lo; rt[f"B{r}"] = hi; rt[f"C{r}"] = val
note(rt, f"A{5+len(tiers)+1}", "Formula: $1 x Billed Tier Value x Frequency Multiplier x Bank Feeds Multiplier x Locations Multiplier")

# Recording frequency multiplier
base = 5+len(tiers)+3
rt[f"A{base}"] = "Recording Frequency Multiplier"
rt[f"A{base}"].font = Font(bold=True, name="Arial")
header(rt, f"A{base+1}", "Frequency")
header(rt, f"B{base+1}", "Multiplier")
freqs = [("Monthly",1),("Bi-weekly",1.5),("Weekly",2)]
for i,(k,v) in enumerate(freqs):
    rt[f"A{base+2+i}"] = k; rt[f"B{base+2+i}"] = v

# Bank feeds multiplier
base2 = base+2+len(freqs)+1
rt[f"A{base2}"] = "Bank Feeds Multiplier"
rt[f"A{base2}"].font = Font(bold=True, name="Arial")
header(rt, f"A{base2+1}", "Bank Feeds")
header(rt, f"B{base2+1}", "Multiplier")
bf = [("Yes",1),("No / manual",2.25)]
for i,(k,v) in enumerate(bf):
    rt[f"A{base2+2+i}"] = k; rt[f"B{base2+2+i}"] = v

# Locations multiplier
base3 = base2+2+len(bf)+1
rt[f"A{base3}"] = "Locations Multiplier"
rt[f"A{base3}"].font = Font(bold=True, name="Arial")
header(rt, f"A{base3+1}", "Number of Locations")
header(rt, f"B{base3+1}", "Multiplier")
locs = [(1,1),(2,1.25),(3,1.75),(4,2),(5,2.25),(6,2.5),(7,2.75)]
for i,(k,v) in enumerate(locs):
    rt[f"A{base3+2+i}"] = k; rt[f"B{base3+2+i}"] = v

# Reconcile accounts tier
base4 = base3+2+len(locs)+1
rt[f"A{base4}"] = "Reconcile Bank/CC/PayPal Accounts — Tier Table"
rt[f"A{base4}"].font = Font(bold=True, name="Arial")
header(rt, f"A{base4+1}", "Lower Bound (>=)")
header(rt, f"B{base4+1}", "Upper Bound")
header(rt, f"C{base4+1}", "Billed Tier Value")
acct_tiers = [(1,5,5),(6,10,10),(11,15,15),(16,25,25),(26,30,30)]
for i,(lo,hi,val) in enumerate(acct_tiers):
    r = base4+2+i
    rt[f"A{r}"] = lo; rt[f"B{r}"] = hi; rt[f"C{r}"] = val
note(rt, f"A{base4+2+len(acct_tiers)+1}", "Formula: $18.50 x Billed Tier Value")

# Advisory tiers
base5 = base4+2+len(acct_tiers)+3
rt[f"A{base5}"] = "Accounting Advisory Services — Flat Tiers"
rt[f"A{base5}"].font = Font(bold=True, name="Arial")
header(rt, f"A{base5+1}", "Company Size")
header(rt, f"B{base5+1}", "Monthly Fee")
adv = [("Small",100),("Medium",150),("Large",200)]
for i,(k,v) in enumerate(adv):
    rt[f"A{base5+2+i}"] = k; rt[f"B{base5+2+i}"] = v

# Sales tax filing frequency divisor
base6 = base5+2+len(adv)+1
rt[f"A{base6}"] = "Sales Tax Filing — Monthly Equivalent Divisor"
rt[f"A{base6}"].font = Font(bold=True, name="Arial")
header(rt, f"A{base6+1}", "Filing Frequency")
header(rt, f"B{base6+1}", "Divisor")
stf = [("Monthly",1),("Quarterly",3),("Annual",12)]
for i,(k,v) in enumerate(stf):
    rt[f"A{base6+2+i}"] = k; rt[f"B{base6+2+i}"] = v
note(rt, f"A{base6+2+len(stf)+1}", "Formula: $60 x Number of States / Divisor")

# Fixed/manual overrides (not on the old matrix — Julia's flat rates)
base7 = base6+2+len(stf)+3
rt[f"A{base7}"] = "Manual Overrides (superseding legacy matrix — confirmed 2026-07-07/08)"
rt[f"A{base7}"].font = Font(bold=True, name="Arial")
rt[f"A{base7+1}"] = "Tax Prep (partnership, Form 1065) default annual fee"
rt[f"B{base7+1}"] = 650
rt[f"A{base7+2}"] = "Tax Prep (S-corp, Form 1120S) default annual fee"
rt[f"B{base7+2}"] = 750
rt[f"A{base7+3}"] = "Tax Prep (C-corp, Form 1120) default annual fee (assumed)"
rt[f"B{base7+3}"] = 780
rt[f"A{base7+4}"] = "1099 Filing, per contractor"
rt[f"B{base7+4}"] = 25
rt[f"A{base7+5}"] = "Financial Statements (P&L + Balance Sheet), flat monthly"
rt[f"B{base7+5}"] = 16
rt[f"A{base7+6}"] = "Owner Payroll, solo S-corp owner (complimentary)"
rt[f"B{base7+6}"] = 0

for col,w in zip("ABCD",[46,16,18,14]):
    rt.column_dimensions[col].width = w

wb.save("/tmp/pricing_wip.xlsx")
print("rate table base rows:", base, base2, base3, base4, base5, base6, base7)
